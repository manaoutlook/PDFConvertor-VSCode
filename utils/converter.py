import logging
import tempfile
import pandas as pd
import tabula
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def clean_amount(amount_str):
    """Clean and format amount strings"""
    if pd.isna(amount_str):
        return ''
    # Remove currency symbols and cleanup
    amount_str = str(amount_str).replace('$', '').replace(',', '').strip()
    # Handle brackets for negative numbers
    if '(' in amount_str and ')' in amount_str:
        amount_str = '-' + amount_str.replace('(', '').replace(')', '')
    return amount_str

def parse_date(date_str):
    """Parse date string from ANZ statement format"""
    try:
        # Clean the date string
        date_str = str(date_str).strip()
        # If it's just day and month (e.g., "26 APR"), add current year
        if len(date_str.split()) == 2:
            current_year = datetime.now().year
            date_str = f"{date_str} {str(current_year)[-2:]}"  # Add last 2 digits of current year
        return datetime.strptime(date_str, '%d %b %y')
    except (ValueError, TypeError) as e:
        logging.debug(f"Failed to parse date: {date_str}, error: {str(e)}")
        return None

def process_transaction_rows(table):
    """Process rows and handle multi-line transactions"""
    processed_data = []
    current_transaction = None

    for _, row in table.iterrows():
        # Clean row values
        row_values = [str(val).strip() if not pd.isna(val) else '' for val in row]

        # Check if this is a continuation line (no valid date)
        date = parse_date(row_values[0])

        if date:
            # If we have a pending transaction, add it
            if current_transaction:
                processed_data.append(current_transaction)

            # Start new transaction
            current_transaction = {
                'Date': date.strftime('%d %b'),
                'Transaction Details': row_values[1],
                'Withdrawals ($)': clean_amount(row_values[2]),
                'Deposits ($)': clean_amount(row_values[3]),
                'Balance ($)': clean_amount(row_values[4]) if len(row_values) > 4 else ''
            }
        else:
            # This is a continuation line
            if current_transaction and row_values[1].strip():
                # Append the additional details to the current transaction
                current_transaction['Transaction Details'] += f"\n{row_values[1].strip()}"

                # If this line has monetary values, use them (they might be vertically centered)
                if clean_amount(row_values[2]):
                    current_transaction['Withdrawals ($)'] = clean_amount(row_values[2])
                if clean_amount(row_values[3]):
                    current_transaction['Deposits ($)'] = clean_amount(row_values[3])
                if clean_amount(row_values[4]):
                    current_transaction['Balance ($)'] = clean_amount(row_values[4])

    # Add the last transaction if pending
    if current_transaction:
        processed_data.append(current_transaction)

    return processed_data

def convert_pdf_to_data(pdf_path: str):
    """Extract data from PDF bank statement and return as list of dictionaries"""
    try:
        logging.info(f"Starting data extraction from {pdf_path}")

        # File validation
        if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
            logging.error("PDF file not found or empty")
            return None

        # Check Java availability
        try:
            import subprocess
            subprocess.run(['java', '-version'], capture_output=True, check=True)
        except Exception as e:
            logging.error(f"Java not available: {str(e)}")
            return None

        # Extract tables with specific settings for ANZ statements
        try:
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                guess=True,
                lattice=False,
                stream=True,
                pandas_options={'header': None},
                java_options=['-Dfile.encoding=UTF8', '-Djava.awt.headless=true']
            )
            logging.info(f"Successfully extracted {len(tables)} tables from PDF")
        except Exception as e:
            logging.error(f"Error during PDF table extraction: {str(e)}")
            return None

        all_transactions = []

        # Process each table
        for idx, table in enumerate(tables):
            logging.debug(f"Processing table {idx+1}, shape: {table.shape}")
            if len(table.columns) >= 4:  # Ensure table has enough columns
                table.columns = range(len(table.columns))

                # Process transactions with multi-line handling
                transactions = process_transaction_rows(table)
                all_transactions.extend(transactions)

        if not all_transactions:
            logging.error("No valid transactions found after processing")
            return None

        logging.info(f"Successfully processed {len(all_transactions)} transactions")
        return all_transactions

    except Exception as e:
        logging.error(f"Error in data extraction: {str(e)}")
        return None

def convert_pdf(pdf_path: str, output_format: str = 'excel'):
    """Convert PDF bank statement to Excel/CSV using tabula-py"""
    try:
        # Extract data using the improved processing logic
        processed_data = convert_pdf_to_data(pdf_path)

        if not processed_data:
            return None

        # Convert to DataFrame
        df = pd.DataFrame(processed_data)

        # Create output file
        temp_file = tempfile.NamedTemporaryFile(delete=False)

        if output_format == 'excel':
            output_path = f"{temp_file.name}.xlsx"
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Transactions')
                workbook = writer.book
                worksheet = writer.sheets['Transactions']

                # Format headers
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                header_alignment = Alignment(horizontal='center')

                for col in range(len(df.columns)):
                    cell = worksheet.cell(row=1, column=col+1)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Auto-adjust column widths
                for idx, column in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

                # Set wrap text for transaction details column
                for cell in worksheet['B']:
                    cell.alignment = Alignment(wrapText=True)
        else:
            output_path = f"{temp_file.name}.csv"
            df.to_csv(output_path, index=False)

        logging.info(f"Successfully created output file: {output_path}")
        return output_path

    except Exception as e:
        logging.error(f"Error in conversion: {str(e)}")
        return None